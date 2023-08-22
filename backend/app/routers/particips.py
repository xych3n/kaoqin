import dataclasses
import re
from typing import Annotated, Any, Optional, Sequence

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from fastapi import APIRouter, Depends, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.orm import Session
from starlette.status import HTTP_400_BAD_REQUEST, HTTP_404_NOT_FOUND, HTTP_409_CONFLICT

from .. import crud, schemas
from ..deps import get_db, get_current_user, get_current_superuser
from ..models import *


router = APIRouter()


@router.get("/", response_model=schemas.Pagination[schemas.Particip])
def read_particips(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    *,
    skip: int = 0,
    limit: int = 100,
    filter: str = "",
) -> Any:
    db_particips = crud.particip.get_multi(db, skip=skip, limit=limit, filter=filter)
    return {
        "total": crud.particip.count(db, filter=filter),
        "list": [dataclasses.asdict(db_particip)
                 | {"activity": db_particip.activity, "participant": db_particip.participant}
                 for db_particip in db_particips]
    }


@router.post("/", response_model=schemas.Particip)
def create_particip(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    *,
    particip_in: schemas.ParticipCreate,
) -> Any:
    db_particip = crud.particip.get(db, activity_id=particip_in.activity_id, user_id=particip_in.user_id)
    if db_particip is not None:
        raise HTTPException(HTTP_409_CONFLICT, "particip existed")
    db_particip = crud.particip.create(db, obj_in=particip_in)
    return dataclasses.asdict(db_particip) | {"activity": db_particip.activity, "participant": db_particip.participant}


@router.delete("/", response_model=schemas.Particip)
def remove_particips(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    *,
    activity_id: int,
    user_id: int,
) -> Any:
    db_particip = db.get(Particip, {"activity_id": activity_id, "user_id": user_id})
    if db_particip is None:
        raise HTTPException(HTTP_404_NOT_FOUND, "particip not found")
    removed_particip = dataclasses.asdict(db_particip) | {"activity": db_particip.activity, "participant": db_particip.participant}
    crud.particip.remove(db, db_obj=db_particip)
    return removed_particip


@router.get("/download", response_class=FileResponse)
async def download_particips(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Any:
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("次数统计")
    ws.append(["学号", "姓名", "讲座次数", "研会次数"])
    if current_user.is_superuser:
        db_users = db.scalars(select(User)).all()
    else:
        db_users = db.scalars(select(User).where(User.id == current_user.id)).all()
    for db_user in db_users:
        involvements = db_user.involvements
        ws.append([db_user.student_number, db_user.name, involvements["学术讲座"], involvements["研会活动"]])
    ws = wb.create_sheet("记录明细")
    if current_user.is_superuser:
        ws.append(["学号", "姓名", "活动名称", "活动日期", "活动类型", "计入次数"])
        db_particips = db.scalars(
            select(Particip)
                .join(Particip.participant)
                .join(Particip.activity)
                .order_by(User.id, Activity.category, Activity.date.desc())
        ).all()
        for db_particip in db_particips:
            if db_particip.activity.category == ActivityCategory.ASSN and db_particip.activity.date <= datetime.date(2022, 9, 1):
                continue
            ws.append([
                db_particip.participant.student_number,
                db_particip.participant.name,
                db_particip.activity.title,
                db_particip.activity.date,
                db_particip.activity.category.value,
                db_particip.involvement,
            ])
    else:
        ws.append(["活动名称", "活动日期", "活动类型", "计入次数"])
        db_particips = db.scalars(
            select(Particip)
                .join(Particip.participant)
                .join(Particip.activity)
                .where(User.id == current_user.id)
                .order_by(Activity.category, Activity.date.desc())
        ).all()
        for db_particip in db_particips:
            if db_particip.activity.category == ActivityCategory.ASSN and db_particip.activity.date <= datetime.date(2022, 9, 1):
                continue
            ws.append([
                db_particip.activity.title,
                db_particip.activity.date,
                db_particip.activity.category.value,
                db_particip.involvement,
            ])
    filename = f"{current_user.student_number}.xlsx"
    wb.save(filename)
    return filename


def _validate_student_number(db: Session, student_number: Any) -> Optional[User]:
    if not isinstance(student_number, str):
        try:
            student_number = str(int(student_number))
        except:
            return
    if re.match(r"\d{11}", student_number):
        db_user = db.scalar(select(User).where(User.student_number == student_number))
        return db_user


def _validate_involvement(involvement: Any) -> Optional[int]:
    if isinstance(involvement, (str, int)):
        try:
            return int(involvement)
        except:
            pass


@router.post("/upload", response_model=Sequence[Particip])
def upload_particips(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser),
    *,
    title: Annotated[str, Form()],
    date: Annotated[datetime.date, Form()],
    category: Annotated[ActivityCategory, Form()],
    file: UploadFile,
) -> Any:
    db_activity = Activity(
        title=title,
        date=date,
        category=category,
    )
    try:
        db.add(db_activity)
        db.commit()
        db.refresh(db_activity)
    except:
        raise HTTPException(HTTP_400_BAD_REQUEST, "unable to insert activity")
    if file.filename is None:
        raise HTTPException(HTTP_400_BAD_REQUEST, "file has no filename?")
    db_particips = []
    if file.filename.endswith(".xlsx"):
        ws = openpyxl.load_workbook(file.file).active
        if ws is None:
            raise HTTPException(HTTP_400_BAD_REQUEST, "no active sheet in the file")
        assert isinstance(ws, Worksheet)
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) != 3:
                raise HTTPException(HTTP_400_BAD_REQUEST, f"column missing on line {i}")
            student_number, _, involvement = row
            db_user = _validate_student_number(db, student_number)
            if db_user is None:
                raise HTTPException(HTTP_400_BAD_REQUEST, f"illegal student number found on line {i}")
            involvement = _validate_involvement(involvement)
            if involvement is None:
                raise HTTPException(HTTP_400_BAD_REQUEST, f"illegal involvement found on line {i}")
            db_particips.append(Particip(
                user_id=db_user.id,
                activity_id=db_activity.id,
                involvement=involvement,
            ))
    else:
        raise HTTPException(HTTP_400_BAD_REQUEST, "unsupported file extension")
    db.add_all(db_particips)
    db.commit()
    return db_particips
